import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Literal
import zipfile

from bs4 import BeautifulSoup
from odf.opendocument import load as load_odf
from odf.table import CoveredTableCell, Table, TableCell, TableRow
from odf.teletype import extractText
from openpyxl import load_workbook
import xlrd

NormalisedFormat = Literal["csv", "html", "xls", "xlsx", "xlsm", "ods"]

ZIP_MAGIC = b"PK\x03\x04"
OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
PDF_MAGIC = b"%PDF-"

TABULAR_EXTENSIONS = {".csv", ".xls", ".xlsx", ".xlsm", ".ods"}
SKIP_EXTENSIONS = {".pdf", ".doc", ".docx", ".ppt", ".pptx"}
ODS_TRAILING_ROW_REPEAT_LIMIT = 10000


@dataclass(frozen=True)
class NormalisedSheet:
    name: str
    rows: list[list[str]]


def normalise(data: bytes, filename: str) -> list[NormalisedSheet]:
    file_format = detect_file_format(data, filename)
    if file_format is None:
        return []
    if file_format == "csv":
        return [NormalisedSheet(name="default", rows=read_csv_rows(data))]
    if file_format == "html":
        return read_html_sheets(data)
    if file_format == "xls":
        try:
            return read_xls_sheets(data)
        except Exception:
            return []
    if file_format in {"xlsx", "xlsm"}:
        try:
            return read_openxml_sheets(data)
        except zipfile.BadZipFile:
            return []
    if file_format == "ods":
        try:
            return read_ods_sheets(data)
        except zipfile.BadZipFile:
            return []
    return []


def detect_file_format(data: bytes, filename: str) -> NormalisedFormat | None:
    suffix = Path(filename).suffix.lower()
    text_prefix = strip_leading_markup_whitespace(data)

    if data.startswith(PDF_MAGIC):
        return None
    if text_prefix.startswith((b"<", b"<!")):
        return "html"
    if data.startswith(OLE_MAGIC):
        if is_probably_csv_bytes(data):
            return "csv"
        return "xls"
    if data.startswith(ZIP_MAGIC):
        zip_format = detect_zip_format(data)
        if zip_format is not None:
            return zip_format
        if suffix == ".xlsx":
            return "xlsx"
        if suffix == ".xlsm":
            return "xlsm"
        if suffix == ".ods":
            return "ods"
        if is_probably_csv_bytes(data):
            return "csv"
        return None
    if suffix in SKIP_EXTENSIONS:
        return None
    if is_probably_csv_bytes(data):
        return "csv"
    if suffix in TABULAR_EXTENSIONS:
        if suffix == ".xls":
            return "xls"
        if suffix == ".xlsx":
            return "xlsx"
        if suffix == ".xlsm":
            return "xlsm"
        if suffix == ".ods":
            return "ods"
        return "csv"
    return None


def detect_zip_format(data: bytes) -> NormalisedFormat | None:
    import zipfile

    try:
        with zipfile.ZipFile(BytesIO(data)) as archive:
            names = set(archive.namelist())
    except zipfile.BadZipFile:
        return None

    if "mimetype" in names:
        with zipfile.ZipFile(BytesIO(data)) as archive:
            mimetype = archive.read("mimetype")
        if mimetype.startswith(b"application/vnd.oasis.opendocument.spreadsheet"):
            return "ods"

    if "xl/workbook.xml" in names:
        if "xl/vbaProject.bin" in names:
            return "xlsm"
        return "xlsx"

    if "word/document.xml" in names or "ppt/presentation.xml" in names:
        return None

    return None


def strip_leading_markup_whitespace(data: bytes) -> bytes:
    if data.startswith(b"\xef\xbb\xbf"):
        data = data[3:]
    return data.lstrip()


def is_probably_csv_bytes(data: bytes, chunk_size: int = 4096) -> bool:
    chunk = data[:chunk_size]
    if not chunk:
        return False
    if b"\x00" in chunk:
        return False
    for encoding in ("utf-8-sig", "cp1252", "latin1"):
        try:
            text = chunk.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        return False

    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return False
    sample = "\n".join(lines[:10])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = None

    if dialect is not None:
        delimiter = getattr(dialect, "delimiter", None)
        if delimiter is not None and any(delimiter in line for line in lines[:3]):
            return True
    return False


def read_csv_rows(data: bytes) -> list[list[str]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp1252", "latin1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    else:
        if last_error is not None:
            text = data.decode("utf-8-sig", errors="replace")
        else:
            return []

    reader = csv.reader(StringIO(text, newline=""))
    return [normalise_row(row) for row in reader]


def read_html_sheets(data: bytes) -> list[NormalisedSheet]:
    text = data.decode("utf-8", errors="replace")
    soup = BeautifulSoup(text, "html.parser")
    sheets: list[NormalisedSheet] = []
    for index, table in enumerate(soup.find_all("table"), start=1):
        heading = table.find_all_previous(["h2", "h3"], limit=1)
        name = heading[0].get_text(" ", strip=True) if heading else f"table-{index}"
        rows: list[list[str]] = []
        for tr in table.find_all("tr"):
            cells = tr.find_all(["td", "th"])
            rows.append([cell.get_text(" ", strip=True) for cell in cells])
        sheets.append(NormalisedSheet(name=name or f"table-{index}", rows=rows))
    return sheets


def read_xls_sheets(data: bytes) -> list[NormalisedSheet]:
    workbook = xlrd.open_workbook(file_contents=data, formatting_info=True)
    sheets = []
    for sheet in workbook.sheets():
        rows = [[""] * sheet.ncols for _ in range(sheet.nrows)]
        for row_index in range(sheet.nrows):
            for col_index in range(sheet.ncols):
                rows[row_index][col_index] = cell_to_string(sheet.cell_value(row_index, col_index))
        for row_start, row_end, col_start, col_end in sheet.merged_cells:
            value = rows[row_start][col_start]
            for row_index in range(row_start, row_end):
                for col_index in range(col_start, col_end):
                    rows[row_index][col_index] = value
        sheets.append(NormalisedSheet(name=sheet.name, rows=rows))
    return sheets


def read_openxml_sheets(data: bytes) -> list[NormalisedSheet]:
    workbook = load_workbook(BytesIO(data), data_only=True)
    sheets = []
    for worksheet in workbook.worksheets:
        expand_openxml_merged_cells(worksheet)
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append([cell_to_string(value) for value in row])
        sheets.append(NormalisedSheet(name=worksheet.title, rows=rows))
    workbook.close()
    return sheets


def expand_openxml_merged_cells(worksheet) -> None:
    for merged_range in list(worksheet.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        value = worksheet.cell(row=min_row, column=min_col).value
        worksheet.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                worksheet.cell(row=row, column=col).value = value


def read_ods_sheets(data: bytes) -> list[NormalisedSheet]:
    document = load_odf(BytesIO(data))
    sheets = []
    for table in document.spreadsheet.getElementsByType(Table):
        name = table.getAttribute("name") or "Sheet1"
        rows = read_ods_table_rows(table)
        sheets.append(NormalisedSheet(name=name, rows=rows))
    return sheets


def read_ods_table_rows(table: Table) -> list[list[str]]:
    rows: list[list[str]] = []
    merge_map: dict[tuple[int, int], str] = {}
    row_index = 0
    for row in table.getElementsByType(TableRow):
        repeated_rows = int(row.getAttribute("numberrowsrepeated") or 1)
        base_row = build_ods_row(row, row_index, merge_map)
        if repeated_rows > ODS_TRAILING_ROW_REPEAT_LIMIT and is_blank_row(base_row):
            # Real GOV.UK ODS files often encode sheet padding as a final repeated blank
            # row block extending to the spreadsheet row limit. Materialising that would
            # hang the crawler without adding any data.
            break
        for _ in range(repeated_rows):
            rows.append(list(base_row))
            row_index += 1
    return rows


def build_ods_row(row: TableRow, row_index: int, merge_map: dict[tuple[int, int], str]) -> list[str]:
    values: list[str] = []
    col_index = 0
    for cell in row.childNodes:
        if not hasattr(cell, "qname"):
            continue
        if cell.qname == CoveredTableCell().qname:
            col_index += int(cell.getAttribute("numbercolumnsrepeated") or 1)
            continue
        if cell.qname != TableCell().qname:
            continue

        repeated_columns = int(cell.getAttribute("numbercolumnsrepeated") or 1)
        column_span = int(cell.getAttribute("numbercolumnsspanned") or 1)
        row_span = int(cell.getAttribute("numberrowsspanned") or 1)
        value = ods_cell_to_string(cell)

        if value == "" and column_span == 1 and row_span == 1:
            col_index += repeated_columns
            continue

        while len(values) < col_index:
            values.append(merge_map.pop((row_index, len(values)), ""))

        for _ in range(repeated_columns):
            values.append(value)
            if column_span > 1 or row_span > 1:
                # The origin cell is already appended directly; only store covered cells.
                for row_offset in range(row_span):
                    for col_offset in range(column_span):
                        if row_offset == 0 and col_offset == 0:
                            continue
                        target = (row_index + row_offset, col_index + col_offset)
                        merge_map[target] = value
            col_index += 1

    while (row_index, len(values)) in merge_map:
        values.append(merge_map.pop((row_index, len(values)), ""))
    return values


def is_blank_row(row: list[str]) -> bool:
    return all(cell == "" for cell in row)


def ods_cell_to_string(cell: TableCell) -> str:
    value = cell.getAttribute("stringvalue")
    if value is not None:
        return str(value)
    value = cell.getAttribute("value")
    if value is not None:
        return str(value)
    date_value = cell.getAttribute("datevalue")
    if date_value is not None:
        return str(date_value)
    time_value = cell.getAttribute("timevalue")
    if time_value is not None:
        return str(time_value)
    return extractText(cell).replace("\r", "").strip()


def normalise_row(row: list[str]) -> list[str]:
    return [cell_to_string(value) for value in row]


def cell_to_string(value) -> str:
    if value is None:
        return ""
    return str(value)
